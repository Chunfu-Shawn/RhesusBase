import openpyxl
from openpyxl.styles import Font
import numpy as np
import pandas as pd
import osha_utils as su
import argparse
from Bio import SeqIO, Seq


def sga(seq1_name, seq1, seq2_name, seq2):
    # Class alignment
    align = su.Alignment(seq1_name, 
                         seq1, 
                         seq2_name, 
                         seq2)

    # dynamic programming
    G, H = su.dp_algorithm(align.seq1_seq, 
                           align.seq2_seq)

    # find best alignment
    max_score, start = su.opt_solution(G,
                                       align.seq1_seq,
                                       align.seq2_seq)

    # backtracing
    where = su.backtracing(H,
                           align.seq2_seq, 
                           start)

    # count gaps
    align.num_gap = su.count_gap(where)

    # target range
    align.target_start, align.target_end, align.target_seq = su.get_target_seq(align.seq1_seq, 
                                                                               where)

    # solve mismatch & predictive_index
    align.mismatch_type, align.mismatch, align.predictive_index = su.get_mis_pred(align.seq1_seq,
                                                                                  align.seq2_seq,
                                                                                  where)
    align.revise_mismatch_type()

    return align


def main(argsv):
    parser = argparse.ArgumentParser(description="Semi-global alignment for siRNA. (OSHA)")
    parser.add_argument("--subject", type=str, help="Subject fasta")
    parser.add_argument("--query", type=str, help="Query xlsx")
    parser.add_argument("--output", type=str, help="Result file")
    args = parser.parse_args()

    #### Parse fasta
    subject = SeqIO.parse(open(args.subject), 'fasta')
    qwb = openpyxl.load_workbook(args.query)
    qws = qwb.active

    #### Open excel
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    # Define a Font object with the desired font properties
    font = Font(name='Courier New', size=12, bold=True, italic=False)
    # Set the font for all cells in the worksheet
    ws.font = font
    header = ['AS_name',
              'AS_seq',
              'Subject_name',
              'Subject_seq',
              'Gap_number',
              'Target_start',
              'Target_end',
              'Target_seq 5-3',
              'Mismatch_type',
              'Mismatch',
              'Predictive_index']
    ws.append(header)

    for q in qws.iter_rows():
        for s in subject:
            q_id = q[0].value
            q_seq = q[1].value

            # Remove special characters, replace U with T
            q_seq = Seq.Seq(su.remove_special_characters(q_seq).replace('T','U'))
            s.seq = Seq.Seq(su.remove_special_characters(str(s.seq)).replace('T','U'))

            print("Processing: query name: {}, query seq: {}, subject name: {}, subject_seq: {}".format(q_id, q_seq, s.id, s.seq))
            n = len(s.seq)
            m = len(q_seq)
            if  n < m: continue

            # perform sga
            align = sga(s.id,
                        s.seq,
                        q_id,
                        q_seq)

            # Output
            print(str(align.seq1_name),
                  str(align.seq1_seq),
                  str(align.seq2_name),
                  str(align.seq2_seq),
                  align.num_gap,
                  align.target_start,
                  align.target_end,
                  str(align.target_seq),
                  align.mismatch_type,
                  align.mismatch,
                  align.predictive_index)

            # save result
            ws.append([align.seq1_name,
                              str(align.seq1_seq),
                              align.seq2_name,
                              str(align.seq2_seq),
                              align.num_gap,
                              align.target_start,
                              align.target_end,
                              str(align.target_seq),
                              align.mismatch_type,
                              align.mismatch,
                              align.predictive_index])


    wb.save(args.output)


def main_deprecated(argsv):
    seq1_name = 'subject'
    seq1_seq = "ATTAGACCTG"
    '''
    seq1 = "GATGGTTGTCTATTAACTTGTTCAAAAAAGTATCAGGAGTTGTCAAGGCAGAGAAGAGAGTGTTTGCAAA" \
           "AGGGGGAAAGTAGTTTGCTGCCTCTTTAAGACTAGGACTGAGAGAAAGAAGAGGAGAGAGAAAGAAAGGG" \
           "AGAGAAGTTTGAGCCCCAGGCTTAAGCCTTTCCAAAAAATAATAATAACAATCATCGGCGGCGGCAGGAT" \
           "CGGCCAGAGGAGGAGGGAAGCGCTTTTTTTGATCCTGATTCCAGTTTGCCTCTCTCTTTTTTTCCCCCAA" \
           "ATTATTCTTCGCCTGATTTTCCTCGCGGAGCCCTGCGCTCCCGACACCCCCGCCCGCCTCCCCTCCTCCT" \
           "CTCCCCCCGCCCGCGGGCCCCCCAAAGTCCCGGCCGGGCCGAGGGTCGGCGGCCGCCGGCGGGCCGGGCC" \
           "CGCGCACAGCGCCCGCATGTACAACATGATGGAGACGGAGCTGAAGCCGCCGGGCCCGCAGCAAACTTCG" \
           "GGGGGCGGCGGCGGCAACTCCACCGCGGCGGCGGCCGGCGGCAACCAGAAAAACAGCCCGGACCGCGTCA" \
           "AGCGGCCCATGAATGCCTTCATGGTGTGGTCCCGCGGGCAGCGGCGCAAGATGGCCCAGGAGAACCCCAA" \
           "GATGCACAACTCGGAGATCAGCAAGCGCCTGGGCGCCGAGTGGAAACTTTTGTCGGAGACGGAGAAGCGG" \
           "CCGTTCATCGACGAGGCTAAGCGGCTGCGAGCGCTGCACATGAAGGAGCACCCGGATTATAAATACCGGC" \
           "CCCGGCGGAAAACCAAGACGCTCATGAAGAAGGATAAGTACACGCTGCCCGGCGGGCTGCTGGCCCCCGG" \
           "CGGCAATAGCATGGCGAGCGGGGTCGGGGTGGGCGCCGGCCTGGGCGCGGGCGTGAACCAGCGCATGGAC" \
           "AGTTACGCGCACATGAACGGCTGGAGCAACGGCAGCTACAGCATGATGCAGGACCAGCTGGGCTACCCGC" \
           "AGCACCCGGGCCTCAATGCGCACGGCGCAGCGCAGATGCAGCCCATGCACCGCTACGACGTGAGCGCCCT" \
           "GCAGTACAACTCCATGACCAGCTCGCAGACCTACATGAACGGCTCGCCCACCTACAGCATGTCCTACTCG" \
           "CAGCAGGGCACCCCTGGCATGGCTCTTGGCTCCATGGGTTCGGTGGTCAAGTCCGAGGCCAGCTCCAGCC" \
           "CCCCTGTGGTTACCTCTTCCTCCCACTCCAGGGCGCCCTGCCAGGCCGGGGACCTCCGGGACATGATCAG" \
           "CATGTATCTCCCCGGCGCCGAGGTGCCGGAACCCGCCGCCCCCAGCAGACTTCACATGTCCCAGCACTAC" \
           "CAGAGCGGCCCGGTGCCCGGCACGGCCATTAACGGCACACTGCCCCTCTCACACATGTGAGGGCCGGACA" \
           "GCGAACTGGAGGGGGGAGAAATTTTCAAAGAAAAACGAGGGAAATGGGAGGGGTGCAAAAGAGGAGAGTA" \
           "AGAAACAGCATGGAGAAAACCCGGTACGCTCAAAAAGAAAAAGGAAAAAAAAAAATCCCATCACCCACAG" \
           "CAAATGACAGCTGCAAAAGAGAACACCAATCCCATCCACACTCACGCAAAAACCGCGATGCCGACAAGAA" \
           "AACTTTTATGAGAGAGATCCTGGACTTCTTTTTGGGGGACTATTTTTGTACAGAGAAAACCTGGGGAGGG" \
           "TGGGGAGGGCGGGGGAATGGACCTTGTATAGATCTGGAGGAAAGAAAGCTACGAAAAACTTTTTAAAAGT" \
           "TCTAGTGGTACGGTAGGAGCTTTGCAGGAAGTTTGCAAAAGTCTTTACCAATAATATTTAGAGCTAGTCT" \
           "CCAAGCGACGAAAAAAATGTTTTAATATTTGCAAGCAACTTTTGTACAGTATTTATCGAGATAAACATGG" \
           "CAATCAAAATGTCCATTGTTTATAAGCTGAGAATTTGCCAATATTTTTCAAGGAGAGGCTTCTTGCTGAA" \
           "TTTTGATTCTGCAGCTGAAATTTAGGACAGTTGCAAACGTGAAAAGAAGAAAATTATTCAAATTTGGACA" \
           "TTTTAATTGTTTAAAAATTGTACAAAAGGAAAAAATTAGAATAAGTACTGGCGAACCATCTCTGTGGTCT" \
           "TGTTTAAAAAGGGCAAAAGTTTTAGACTGTACTAAATTTTATAACTTACTGTTAAAAGCAAAAATGGCCA" \
           "TGCAGGTTGACACCGTTGGTAATTTATAATAGCTTTTGTTCGATCCCAACTTTCCATTTTGTTCAGATAA" \
           "AAAAAACCATGAAATTACTGTGTTTGAAATATTTTCTTATGGTTTGTAATATTTCTGTAAATTTATTGTG" \
           "ATATTTTAAGGTTTTCCCCCCTTTATTTTCCGTAGTTGTATTTTAAAAGATTCGGCTCTGTATTATTTGA" \
           "ATCAGTCTGCCGAGAATCCATGTATATATTTGAACTAATATCATCCTTATAACAGGTACATTTTCAACTT" \
           "AAGTTTTTACTCCATTATGCACAGTTTGAGATAAATAAATTTTTGAAATATGGACACTGAAA"
       '''
    seq2_name = 'subject'
    seq2_seq = "ATAGCTCTC"
    '''
    seq2 = "CGGCGGGAACAGCATGGCGAG"
    '''
    
    align = sga(seq1_name,
                seq1_seq,
                seq2_name,
                seq2_seq)

    # Output
    print(align.seq1_name,
          align.seq1_seq,
          align.seq2_name,
          align.seq2_seq,
          align.num_gap,
          align.target_start,
          align.target_end,
          align.target_seq,
          align.mismatch_type,
          align.mismatch,
          align.predictive_index)


if __name__ == '__main__':
    import sys

    main(sys.argv)
