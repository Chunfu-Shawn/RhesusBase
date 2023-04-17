import re
import argparse
import openpyxl
from openpyxl.styles import Font
from Bio import SeqIO, Seq


class Sirna:
    def __init__(self, id, start, length_ss, length_as, overhang, length_ss_3, length_as_3, end_with):
        self.id = id
        self.start = start
        self.length_ss = length_ss
        self.length_as = length_as
        self.overhang = overhang
        self.length_ss_3 = length_ss_3
        self.length_as_3 = length_as_3
        self.end_with = end_with
        self.length_ss_m = self.length_ss - self.length_ss_3
        self.length_as_m = self.length_as - self.length_as_3
        if self.overhang == "AS 3'" and self.length_ss == 19 and self.length_as == 21:
            self.length_ss_3 = 0
            self.length_as_3 = 2
        elif self.overhang == "SS 3'":
            self.length_as_3 =0
        elif self.overhang == "Both" and self.length_ss == 20 and self.length_as == 21:
            self.length_ss_3 = 1
            self.length_as_3 = 2


    def reverse_complement(self,seq):
        # Create a dictionary of complement bases
        base_complement = {'A': 'U', 'C': 'G', 'G': 'C', 'T': 'A', 'U': 'A'}
        # Reverse complement the sequence
        rev_seq = seq[::-1]
        rev_comp_seq = ''.join([base_complement.get(base, base) for base in rev_seq])
        return rev_comp_seq

    def sisco(self,seq):
        self.seq_subject = seq[(self.start - 1):(self.start - 1 + self.length_ss)]

        if self.end_with == 'UU':
            self.seq_ss = seq[(self.start - 1):(self.start - 1 + self.length_ss_m)].replace('T', 'U') + 'U' * self.length_ss_3
            self.seq_as = self.reverse_complement(seq[(self.start - 1):(self.start - 1 + self.length_as_m)].replace('T', 'U')) + 'U' * self.length_as_3
            self.seq_ss_3 = 'U' * self.length_ss_3
            self.seq_as_3 = 'U' * self.length_as_3
        elif self.end_with == 'TT':
            self.seq_ss = seq[(self.start - 1):(self.start - 1 + self.length_ss_m)].replace('T', 'U') + 'T' * self.length_ss_3
            self.seq_as = self.reverse_complement(seq[(self.start - 1):(self.start - 1 + self.length_as_m)].replace('T', 'U')) + 'T' * self.length_as_3
            self.seq_ss_3 = 'T' * self.length_ss_3
            self.seq_as_3 = 'T' * self.length_as_3
        elif self.end_with == 'NN':
            self.seq_ss = seq[(self.start - 1):(self.start - 1 + self.length_ss)].replace('T', 'U')
            self.seq_as = self.reverse_complement(seq[(self.start - 1 - self.length_as_3):(self.start - 1 + self.length_as_m)].replace('T', 'U'))
            self.seq_ss_3 = self.seq_ss[-self.length_ss_3:]
            self.seq_as_3 = self.seq_as[-self.length_as_3:]

        if len(self.seq_ss) != self.length_ss:
            self.seq_ss = None
            self.seq_ss_3 = None
            self.seq_ss_3 = None
        if len(self.seq_as) != self.length_as:
            self.seq_as = None
            self.seq_as_3 = None
            self.seq_as_3 = None
        if len(self.seq_subject) != self.length_ss:
            self.seq_subject = None


    def start_end(self,seq):
        # SS
        self.start_ss = self.start if self.start < len(seq) else None
        if self.start_ss == None:
            self.end_ss = None
        elif self.start_ss + self.length_ss - 1 > len(seq):
            self.end_ss = None
        else:
            self.end_ss = self.start_ss + self.length_ss - 1

        # AS
        if self.start_ss == None:
            self.start_as = None
        elif self.start_ss - self.length_as_3 <= 0:
            self.start_as = None
        else:
            self.start_as = self.start_ss - self.length_as_3

        if self.start_as == None:
            self.end_as = None
        elif self.start_as + self.length_as - 1 > len(seq):
            self.end_as = None
        else:
            self.end_as = self.start_as + self.length_as - 1

        # Subject
        self.start_subject = self.start_ss
        self.end_subject = self.end_ss


def remove_special_characters(string):
    string = re.sub(r'[^a-zA-Z]+', '', string)
    return string


def main(argsv):
    parser = argparse.ArgumentParser(description="2. SISCO")
    parser.add_argument("--subject", type=str, help="Subject fasta")
    parser.add_argument("--query", type=str, help="query siRNA excel")
    parser.add_argument("--length_ss", default=19, type=int, help="Length of SS")
    parser.add_argument("--length_as", default=21, type=int, help="Length of AS")
    parser.add_argument("--overhang", default='Both', choices=["SS 3'", "AS 3'", "Both"], type=str, help="Overhang type")
    parser.add_argument("--length_ss_3", default=0, type=int, help="Length of SS 3'")
    parser.add_argument("--length_as_3", default=2, type=int, help="Length of AS 3'")
    parser.add_argument("--end_with", default=19, type=str, choices=['NN','UU','TT'], help="End type")
    parser.add_argument("--with_header", action='store_true', help="If siRNA fle has header")
    parser.add_argument("--output", type=str, help="Result file")
    args = parser.parse_args()

    # Read data
    # siRNA
    swb = openpyxl.load_workbook(args.query)
    sws = swb.active
    # subject
    subject = SeqIO.parse(open(args.subject), 'fasta')
    subject = next(subject)
    s_id = subject.id
    s_seq = subject.seq
    # Remove special characters, replace U with T
    s_seq = remove_special_characters(str(s_seq).replace('T', 'U'))

    # Configure output
    # Create a new workbook and select the active worksheet
    rwb = openpyxl.Workbook()
    rws = rwb.active
    # Define a Font object with the desired font properties
    font = Font(name='Courier New', size=12, bold=True, italic=False)
    # Set the font for all cells in the worksheet
    rws.font = font
    header = ['ID',
              'Overhang_type',
              'End_with',
              'SS_sequence',
              'SS_length',
              'SS_start',
              'SS_end',
              "SS_3'_sequence",
              "SS_3'_length",
              'AS_sequence',
              'AS_length',
              'AS_start',
              'AS_end',
              "AS_3'_sequence",
              "AS_3'_length",
              "Subject_sequence",
              'Subject_start',
              'Subject_end']
    rws.append(header)

    # Run sisco
    row_start = 2 if args.with_header else 1
    for q in sws.iter_rows(min_row=row_start):
        q_id = q[0].value
        start = q[1].value
        # Create Sirna object
        sirna = Sirna(id=q_id,
                      start=start,
                      length_ss=args.length_ss,
                      length_as=args.length_as,
                      overhang=args.overhang,
                      length_ss_3=args.length_ss_3,
                      length_as_3=args.length_as_3,
                      end_with=args.end_with)
        sirna.sisco(s_seq)
        sirna.start_end(s_seq)
        # Save result
        rws.append([sirna.id,
              sirna.overhang,
              sirna.end_with,
              sirna.seq_ss,
              sirna.length_ss,
              sirna.start_ss,
              sirna.end_ss,
              sirna.seq_ss_3,
              sirna.length_ss_3,
              sirna.seq_as,
              sirna.length_as,
              sirna.start_as,
              sirna.end_as,
              sirna.seq_as_3,
              sirna.length_as_3,
              sirna.seq_subject,
              sirna.start_subject,
              sirna.end_subject])

    rwb.save(args.output)


if __name__ == '__main__':
    import sys

    main(sys.argv)
